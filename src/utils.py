import re
import os
import io
import sys
import json
import time
import struct
import piexif
import zipfile
from tqdm import tqdm
from pathlib import Path
from mutagen.mp4 import MP4
from InquirerPy import inquirer
from collections import defaultdict
from PIL import PngImagePlugin, Image
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, LineChart


def ask(question, options):
    if sys.stdin.isatty():
        return inquirer.select(
            message=question,
            choices=options
        ).execute()
    else:
        print(question)
        for i, option in enumerate(options, 1):
            print(f"{i}. {option}")
        while True:
            try:
                choice = int(input("Choose an option: "))
                if 1 <= choice <= len(options):
                    return options[choice - 1]
            except ValueError:
                print("Please enter a valid choice.")

def ask_number(question):
    if sys.stdin.isatty():
        return int(inquirer.number(
            message=question
        ).execute())
    else:
        print(question)
        while True:
            try:
                value = int(input("⮕ "))
                return value
            except ValueError:
                print("Please enter a valid number.")

def get_mp4_duration(path, file_path):
    """Directly reads the duration of an MP4 file by parsing the 'mvhd' box."""
    try:
        with zipfile.ZipFile(path, mode="r") as package:
            with package.open(file_path, "r") as f:
                data = f.read()
                mvhd_pos = data.find(b'mvhd')
                if mvhd_pos == -1:
                    print(f"Error: 'mvhd' not found in {file_path}")
                    return 0
                mvhd_offset = mvhd_pos + 4
                version = data[mvhd_offset]

                if version == 0:
                    time_scale, duration = struct.unpack(">II", data[mvhd_offset + 12: mvhd_offset + 20])
                elif version == 1:
                    time_scale, duration = struct.unpack(">IQ", data[mvhd_offset + 20: mvhd_offset + 32])
                else:
                    print(f"Error: Unknown version of mvhd for {file_path}")
                    return 0
                return duration / time_scale if time_scale > 0 else 0
    except Exception as e:
        print(f"Error with {file_path} : {e}")
        return 0


def generate_excel(per_contact_stats, messages_per_day, hour_distribution, excel_name="analysis.xlsx"):
    base_name = excel_name if excel_name.endswith(".xlsx") else f"{excel_name}.xlsx"
    count = 1
    while os.path.exists(f"Excels/{base_name}"):
        base_name = f"{excel_name}{count}.xlsx"
        if not base_name.endswith(".xlsx"):
            base_name += ".xlsx"
        count += 1

    wb = Workbook()

    # sort contact by msg
    sorted_indexes = sorted(
        range(len(per_contact_stats["Messages sent by you"])),
        key=lambda i: per_contact_stats["Messages sent by you"][i],
        reverse=True
    )
    for key in per_contact_stats:
        per_contact_stats[key] = [per_contact_stats[key][i] for i in sorted_indexes]

    charts_ws = wb.active
    charts_ws.title = "Charts"

    # ===== Global =====
    ws1 = wb.create_sheet("Global")

    keys = list(per_contact_stats.keys())
    length = len(per_contact_stats[keys[0]])

    for col, key in enumerate(keys, 1):
        ws1.cell(row=1, column=col, value=key)

    for row in range(length):
        for col, key in enumerate(keys, 1):
            val = per_contact_stats[key][row]
            cell = ws1.cell(row=row + 2, column=col, value=val)
            if isinstance(val, timedelta):
                cell.value = val.total_seconds() / 86400
                cell.number_format = '[h]:mm:ss'

    total_row = length + 3
    ws1.cell(row=total_row, column=1, value="TOTAL")
    for col, key in enumerate(keys[1:], 2):
        col_letter = get_column_letter(col)
        if "delay" in key.lower() or "voice" in key.lower():
            cell = ws1.cell(row=total_row, column=col,
                            value=f"=AVERAGE({col_letter}2:{col_letter}{length + 1})")
            cell.number_format = '[h]:mm:ss'
        else:
            ws1.cell(row=total_row, column=col,
                     value=f"=SUM({col_letter}2:{col_letter}{length + 1})")

    # ===== Messages per day =====
    ws2 = wb.create_sheet("Messages per day")
    all_dates = sorted(set(datetime.strptime(date, "%d/%m/%Y")
                           for date in messages_per_day.keys()))
    contacts = sorted({c for daily in messages_per_day.values() for c in daily})
    ws2.cell(row=1, column=1, value="Date")
    for col, contact in enumerate(contacts, 2):
        ws2.cell(row=1, column=col, value=contact)
    ws2.cell(row=1, column=len(contacts) + 2, value="TOTAL")

    for row, dt in enumerate(all_dates, 2):
        date_str = dt.strftime("%d/%m/%Y")
        ws2.cell(row=row, column=1, value=date_str)
        total = 0
        for col, contact in enumerate(contacts, 2):
            msg_count = sum(messages_per_day.get(date_str, {}).get(contact, (0, 0)))
            ws2.cell(row=row, column=col, value=msg_count)
            total += msg_count
        ws2.cell(row=row, column=len(contacts) + 2, value=total)

    # ===== Messages per week =====
    ws_week = wb.create_sheet("Messages per week")
    weekly_data = {}
    for dt in all_dates:
        week = dt.strftime("%Y-W%U")
        weekly_data[week] = weekly_data.get(week, 0) + sum(
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[0] +
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[1]
            for c in contacts
        )
    ws_week.cell(row=1, column=1, value="Week")
    ws_week.cell(row=1, column=2, value="Messages")
    for i, (week, count) in enumerate(sorted(weekly_data.items()), start=2):
        ws_week.cell(row=i, column=1, value=week)
        ws_week.cell(row=i, column=2, value=count)

    # ===== Messages per month =====
    ws_month = wb.create_sheet("Messages per month")
    monthly_data = {}
    for dt in all_dates:
        month = dt.strftime("%Y-%m")
        monthly_data[month] = monthly_data.get(month, 0) + sum(
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[0] +
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[1]
            for c in contacts
        )
    ws_month.cell(row=1, column=1, value="Month")
    ws_month.cell(row=1, column=2, value="Messages")
    for i, (month, count) in enumerate(sorted(monthly_data.items()), start=2):
        ws_month.cell(row=i, column=1, value=month)
        ws_month.cell(row=i, column=2, value=count)

    # ===== Messages per year =====
    ws_year = wb.create_sheet("Messages per year")
    yearly_data = {}
    for dt in all_dates:
        year = dt.strftime("%Y")
        yearly_data[year] = yearly_data.get(year, 0) + sum(
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[0] +
            messages_per_day.get(dt.strftime("%d/%m/%Y"), {}).get(c, (0, 0))[1]
            for c in contacts
        )
    ws_year.cell(row=1, column=1, value="Year")
    ws_year.cell(row=1, column=2, value="Messages")
    for i, (year, count) in enumerate(sorted(yearly_data.items()), start=2):
        ws_year.cell(row=i, column=1, value=year)
        ws_year.cell(row=i, column=2, value=count)

    # ===== Cumulative per day =====
    ws3 = wb.create_sheet("Cumulative per day")
    ws3.cell(row=1, column=1, value="Date")
    for col, contact in enumerate(contacts, 2):
        ws3.cell(row=1, column=col, value=contact)
    ws3.cell(row=1, column=len(contacts) + 2, value="TOTAL")

    cumulative = [0] * len(contacts)
    for row, dt in enumerate(all_dates, 2):
        date_str = dt.strftime("%d/%m/%Y")
        ws3.cell(row=row, column=1, value=date_str)
        total = 0
        for col, contact in enumerate(contacts, 2):
            msg_count = sum(messages_per_day.get(date_str, {}).get(contact, (0, 0)))
            cumulative[col - 2] += msg_count
            ws3.cell(row=row, column=col, value=cumulative[col - 2])
            total += cumulative[col - 2]
        ws3.cell(row=row, column=len(contacts) + 2, value=total)

    # ===== Message Activity by Hour =====
    ws_hour = wb.create_sheet("Message Activity by Hour")
    ws_hour.cell(row=1, column=1, value="Hour")
    ws_hour.cell(row=1, column=2, value="Nb msg")
    for i in range(24):
        ws_hour.cell(row=i + 2, column=1, value=i)
        ws_hour.cell(row=i + 2, column=2, value=hour_distribution[i])


    def place_chart(chart, pos):
        chart.height = 12
        chart.width = 21
        charts_ws.add_chart(chart, pos)

    def make_bar(title, xlab, ylab, ws, data_col, cat_col, max_row):
        chart = BarChart()
        chart.title = title
        chart.x_axis.title = xlab
        chart.y_axis.title = ylab
        data = Reference(ws, min_col=data_col, min_row=2, max_row=max_row)
        cats = Reference(ws, min_col=cat_col, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        return chart

    # Cumulative chart
    cum_chart = LineChart()
    cum_chart.title = "Cumulative Messages per Day"
    data = Reference(ws3, min_col=2, max_col=len(contacts) + 1, min_row=1, max_row=len(all_dates) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(all_dates) + 1)
    cum_chart.add_data(data, titles_from_data=True)
    cum_chart.set_categories(cats)
    place_chart(cum_chart, "A1")

    # Delay chart
    delay_chart = BarChart()
    delay_chart.type = "bar"
    delay_chart.title = "Answer Delay Comparison (minutes)"
    delay_chart.y_axis.title = "Contact"
    delay_chart.x_axis.title = "Minutes"
    names, you_delays, contact_delays = [], [], []
    for i in range(min(5, length)):
        names.append(per_contact_stats["Contact"][i])
        if "Your answer delay" in per_contact_stats:
            you_delays.append(-per_contact_stats["Your answer delay"][i].total_seconds() / 60)
            contact_delays.append(per_contact_stats["Contact answer delay"][i].total_seconds() / 60)

    for i, name in enumerate(names, start=2):
        charts_ws.cell(row=i, column=27, value=name)
        if "Your answer delay" in per_contact_stats:
            charts_ws.cell(row=i, column=28, value=you_delays[i - 2])
            charts_ws.cell(row=i, column=29, value=contact_delays[i - 2])
    charts_ws.cell(row=1, column=27, value="Contact")
    charts_ws.cell(row=1, column=28, value="You")
    charts_ws.cell(row=1, column=29, value="Contact")

    data = Reference(charts_ws, min_col=28, max_col=29, min_row=1, max_row=1 + len(names))
    cats = Reference(charts_ws, min_col=27, min_row=2, max_row=1 + len(names))
    delay_chart.add_data(data, titles_from_data=True)
    delay_chart.set_categories(cats)
    place_chart(delay_chart, "N1")

    # place bar
    place_chart(make_bar("Message Activity by Hour", "Hour", "Messages", ws_hour, 2, 1, 25), "A25")
    place_chart(make_bar("Messages per Day", "Date", "Messages", ws2, len(contacts)+2, 1, len(all_dates)+1), "N25")
    place_chart(make_bar("Messages per Week", "Week", "Messages", ws_week, 2, 1, 1 + len(weekly_data)), "A49")
    place_chart(make_bar("Messages per Month", "Month", "Messages", ws_month, 2, 1, 1 + len(monthly_data)), "N49")
    place_chart(make_bar("Messages per Year", "Year", "Messages", ws_year, 2, 1, 1 + len(yearly_data)), "A73")

    # Pie chart
    pie_chart = PieChart()
    pie_chart.title = "Messages per Contact"
    labels = Reference(ws1, min_col=1, min_row=2, max_row=1 + length)
    data = Reference(ws1, min_col=2, min_row=2, max_row=1 + length)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 12
    pie_chart.width = 21
    charts_ws.add_chart(pie_chart, "N73")

    # ===== sheets order =====
    sheet_order = [
        charts_ws.title,
        ws1.title,
        ws3.title,
        ws_hour.title,
        ws2.title,
        ws_week.title,
        ws_month.title,
        ws_year.title,
    ]
    for i, sheet_name in enumerate(sheet_order):
        wb.move_sheet(sheet_name, offset=i - wb.sheetnames.index(sheet_name))

    create_directory("Excels")
    wb.save(f"Excels/{base_name}")
    print(f"Excels/{base_name} was successfully created")

def __add_jpeg_metadata(path, dt, contact = None):
    try:
        exif_dict = piexif.load(path)
        date_str = dt.strftime("%Y:%m:%d %H:%M:%S")
        exif_dict["0th"][piexif.ImageIFD.DateTime] = date_str.encode('utf-8')
        exif_dict["Exif"][piexif.ExifIFD.DateTimeOriginal] = date_str.encode('utf-8')
        exif_dict["Exif"][piexif.ExifIFD.DateTimeDigitized] = date_str.encode('utf-8')
        if contact is not None:
            comment = f"Contact: {contact}"
            exif_dict["0th"][piexif.ImageIFD.Artist] = comment.encode('utf-8')
        exif_bytes = piexif.dump(exif_dict)
        piexif.insert(exif_bytes, path)

        timestamp = time.mktime(dt.timetuple())
        os.utime(path, (timestamp, timestamp))
    except Exception as e:
        print(f"[EXIF ERROR] {path}: {e}")

def __add_png_metadata(path, dt, contact=None):
    try:
        img = Image.open(path)
        meta = PngImagePlugin.PngInfo()
        meta.add_text("Creation Time", dt.strftime("%Y-%m-%d %H:%M:%S"))
        meta.add_text("File Access Date", dt.strftime("%Y-%m-%d %H:%M:%S"))
        meta.add_text("File Inode Change Date", dt.strftime("%Y-%m-%d %H:%M:%S"))
        meta.add_text("File Modify Date", dt.strftime("%Y-%m-%d %H:%M:%S"))
        if contact is not None:
            meta.add_text("Author", f"Contact: {contact}")
        img.save(path, pnginfo=meta)
        
        timestamp = time.mktime(dt.timetuple())
        os.utime(path, (timestamp, timestamp))
    except Exception as e:
        print(f"[PNG ERROR] {path}: {e}")

def __add_mp4_metadata(path, dt):
    try:
        mp4 = MP4(path)
        mp4["\xa9day"] = [dt.strftime("%Y-%m-%d")]
        mp4.save()

        timestamp = time.mktime(dt.timetuple())
        os.utime(path, (timestamp, timestamp))
    except Exception as e:
        print(f"[MP4 ERROR] {path}: {e} — applying fallback.")


def add_metadata(path, dt, ext, contact=None):
    if ext in [".jpg", ".jpeg", ".webp"]:
        __add_jpeg_metadata(path, dt, contact)
    elif ext in [".png"]:
        __add_png_metadata(path, dt, contact)
    elif ext in [".mp4"]:
        __add_mp4_metadata(path, dt)
    else:
        timestamp = time.mktime(dt.timetuple())
        os.utime(path, (timestamp, timestamp))



def generate_merge_template(file_path="merge_map.csv"):
    with open(file_path, "w", encoding="utf-8", newline="") as f:
        f.write("# If your contact has the same username on different platforms, you can list them below\n")
        f.write("# Exemple:\n")
        f.write('# "john_d","john.insta","john.snap","john.wapp"\n')
        f.write('# "lucie_d","lucie.snap","lucie.whatsapp"\n')
        f.write("# \n")
        f.write('"","","",""\n')

def create_directory(nested_directory):
    try:
        os.makedirs(nested_directory)
        print(f"Directory '{nested_directory}' created successfully.")
    except FileExistsError:
        pass
    except PermissionError:
        print(f"Permission denied: Unable to create '{nested_directory}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

class Action:
    def __init__(self, label, func):
        self.label = label
        self.func = func

    def __str__(self):
        return self.label

    def execute(self):
        self.func()